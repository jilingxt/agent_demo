from importlib.util import module_from_spec, spec_from_file_location
from pathlib import Path
from time import sleep

import openpyxl


ROOT = Path(__file__).resolve().parents[1]
GENERATOR = ROOT / "scripts" / "generate_bayesian_statistics_workbook.py"
WORKBOOK = ROOT / "docs" / "statistics" / "bayesian_parameter_collection_template.xlsx"
SHEETS = [
    "开始这里",
    "名词解释",
    "参数总览",
    "原子事实记录",
    "根节点统计",
    "来源表现统计",
    "CPD样本记录",
    "CPD组合统计",
    "抽取质量统计",
    "来源依赖",
    "权威锚定复核",
    "候选参数",
    "模型发布记录",
    "不可学习项",
    "完整示例",
    "人工修改流程",
    "枚举值",
]
REQUIRED_HEADERS = {
    "参数总览": {"参数层", "参数类型", "当前值", "能否根据真实数据修改", "如何构建统计量", "人工修改位置"},
    "原子事实记录": {"样本ID", "来源主张标签", "独立核验标签", "是否纳入统计", "统计分类"},
    "根节点统计": {"当前prior", "真数", "假数", "后验均值", "95%下限", "95%上限", "候选决策"},
    "来源表现统计": {"TP", "FP", "FN", "TN", "指标类型", "后验均值", "有效分母"},
    "CPD样本记录": {"父节点1", "值1", "子节点独立核验标签", "数据用途"},
    "CPD组合统计": {"父节点状态组合", "子节点真数", "子节点假数", "后验均值"},
    "候选参数": {"参数类型", "当前值", "候选值", "训练有效样本", "验证有效样本", "审批状态"},
    "模型发布记录": {"参数哈希", "校准状态", "独立复核状态", "离线审批状态", "隐私检查", "回滚包路径"},
}


def _load_generator():
    spec = spec_from_file_location("statistics_workbook_generator", GENERATOR)
    assert spec and spec.loader
    module = module_from_spec(spec)
    spec.loader.exec_module(module)
    return module


def _headers(sheet):
    return {cell.value for cell in sheet[1] if cell.value}


def test_generator_creates_novice_oriented_parameter_contract(tmp_path):
    generated = tmp_path / WORKBOOK.name
    _load_generator().build_workbook(generated)

    workbook = openpyxl.load_workbook(generated, data_only=False)

    assert workbook.sheetnames == SHEETS
    assert all(sheet.freeze_panes == "A2" for sheet in workbook.worksheets)
    assert all(sheet.auto_filter.ref for sheet in workbook.worksheets)
    assert all(sheet.sheet_view.showGridLines is False for sheet in workbook.worksheets)
    assert workbook["CPD样本记录"].max_row == 1001
    assert workbook["来源依赖"].max_row == 301
    assert workbook["模型发布记录"].max_row == 301
    for sheet_name, expected_headers in REQUIRED_HEADERS.items():
        assert expected_headers <= _headers(workbook[sheet_name])

    overview_text = " ".join(
        str(cell.value or "")
        for row in workbook["参数总览"].iter_rows()
        for cell in row
    )
    for model_id in (
        "conduct_result",
        "property_taking",
        "deception_disposition",
        "public_order",
        "public_safety",
        "status_duty",
    ):
        assert model_id in overview_text
    assert "不能由单个比例手工换算" in overview_text
    assert "禁止自动修改" in overview_text

    start_text = " ".join(str(cell.value or "") for row in workbook["开始这里"].iter_rows() for cell in row)
    assert "Beta后验不是单案后验" in start_text
    assert "未知" in start_text
    assert "同源" in start_text

    formulas = [
        cell.value
        for sheet in workbook.worksheets
        for row in sheet.iter_rows()
        for cell in row
        if cell.data_type == "f"
    ]
    assert any("COUNTIFS" in formula and "原子事实记录" in formula for formula in formulas)
    assert any("BETA.INV" in formula for formula in formulas)
    assert any("未知" in formula and '=""' in formula for formula in formulas)

    warning_columns = {
        "根节点统计": "Q2",
        "来源表现统计": "U2",
        "CPD组合统计": "P2",
    }
    for sheet_name, expected_reference in warning_columns.items():
        sheet = workbook[sheet_name]
        conditional_formulas = [
            formula
            for conditional_range in sheet.conditional_formatting
            for rule in sheet.conditional_formatting[conditional_range]
            for formula in rule.formula
        ]
        assert any(expected_reference in formula for formula in conditional_formulas)
        assert all("A1" not in formula for formula in conditional_formulas)

    validations = [
        validation
        for sheet in workbook.worksheets
        for validation in sheet.data_validations.dataValidation
    ]
    assert validations
    assert {"=TruthLabels", "=ApprovalStatuses", "=PrivacyStatuses", "=SourceMetrics"} <= {
        validation.formula1 for validation in validations
    }


def test_checked_in_workbook_matches_current_models_and_formulas():
    assert WORKBOOK.exists()
    workbook = openpyxl.load_workbook(WORKBOOK, data_only=False)
    assert workbook.sheetnames == SHEETS
    assert all(sheet.freeze_panes == "A2" for sheet in workbook.worksheets)
    assert workbook["原子事实记录"]["R2"].value.startswith("=IF(Q2<>")
    assert "COUNTIFS" in workbook["根节点统计"]["F2"].value
    assert "BETA.INV" in workbook["根节点统计"]["O2"].value
    assert workbook["来源表现统计"]["Q2"].value.startswith("=IF(")


def test_generator_output_is_byte_reproducible(tmp_path):
    output = tmp_path / WORKBOOK.name
    generator = _load_generator()

    generator.build_workbook(output)
    first = output.read_bytes()
    sleep(2.1)
    generator.build_workbook(output)

    assert output.read_bytes() == first
