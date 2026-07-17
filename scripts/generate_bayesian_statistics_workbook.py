from __future__ import annotations

import json
from datetime import datetime, timezone
from pathlib import Path
from xml.etree import ElementTree
from zipfile import ZIP_DEFLATED, ZipFile

from openpyxl import Workbook
from openpyxl.comments import Comment
from openpyxl.formatting.rule import CellIsRule, FormulaRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Protection, Side
from openpyxl.utils import get_column_letter, quote_sheetname
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.workbook.defined_name import DefinedName


ROOT = Path(__file__).resolve().parents[1]
MODEL_DIR = ROOT / "config" / "bayesian_models"
AUTHORITY_RULES = ROOT / "config" / "authority_rules.json"
DEFAULT_OUTPUT = ROOT / "docs" / "statistics" / "bayesian_parameter_collection_template.xlsx"

ATOMIC_LAST_ROW = 2001
SUMMARY_LAST_ROW = 301
CPD_LAST_ROW = 1001

HEADER_FILL = PatternFill("solid", fgColor="173F5F")
SECTION_FILL = PatternFill("solid", fgColor="20639B")
INPUT_FILL = PatternFill("solid", fgColor="FFF2CC")
FORMULA_FILL = PatternFill("solid", fgColor="E2F0D9")
CURRENT_FILL = PatternFill("solid", fgColor="E7E6E6")
WARNING_FILL = PatternFill("solid", fgColor="FCE4D6")
DANGER_FILL = PatternFill("solid", fgColor="F4CCCC")
HEADER_FONT = Font(name="Microsoft YaHei", color="FFFFFF", bold=True)
BODY_FONT = Font(name="Microsoft YaHei", size=10)
TITLE_FONT = Font(name="Microsoft YaHei", size=12, bold=True, color="173F5F")
THIN_BORDER = Border(bottom=Side(style="thin", color="B7C9D3"))

CORE_NAMESPACES = {
    "cp": "http://schemas.openxmlformats.org/package/2006/metadata/core-properties",
    "dc": "http://purl.org/dc/elements/1.1/",
    "dcterms": "http://purl.org/dc/terms/",
    "dcmitype": "http://purl.org/dc/dcmitype/",
    "xsi": "http://www.w3.org/2001/XMLSchema-instance",
}
for prefix, uri in CORE_NAMESPACES.items():
    ElementTree.register_namespace(prefix, uri)


ENUMS = {
    "TruthLabels": ("真值标签", ["真", "假", "未知"]),
    "StanceLabels": ("来源立场", ["真", "假", "未知"]),
    "YesNoValues": ("是否", ["是", "否"]),
    "ReviewStatuses": ("复核状态", ["未开始", "待复核", "一致", "有分歧", "已裁决"]),
    "SourceTypes": (
        "来源类别",
        [
            "statement",
            "evidence_image",
            "report_image",
            "official_record",
            "forensic_report",
            "transaction_record",
            "other",
        ],
    ),
    "SourceRoles": (
        "来源角色",
        [
            "reporting_person",
            "alleged_actor",
            "witness",
            "official",
            "expert",
            "neutral_third_party",
            "unknown",
        ],
    ),
    "SourceMetrics": (
        "来源统计指标",
        ["肯定主张正确率", "否定主张正确率", "灵敏度", "特异度"],
    ),
    "DataSplits": ("数据用途", ["训练", "验证", "测试"]),
    "ParameterDecisions": (
        "参数决策",
        ["不采用", "仅作试点候选", "进入独立验证", "批准发布"],
    ),
    "CalibrationStatuses": (
        "校准状态",
        ["expert_prior_unvalidated", "pilot_uncalibrated", "validated"],
    ),
    "ApprovalStatuses": ("审批状态", ["草稿", "待审批", "已批准", "已拒绝"]),
    "PrivacyStatuses": ("隐私检查", ["未完成", "通过", "不通过"]),
    "AuthorityStatuses": (
        "权威复核状态",
        ["待核验", "有效", "有争议", "已失效", "已替代", "不适用"],
    ),
}


QUALITY_DIMENSION_WEIGHTS = {
    "extraction_quality": 0.18,
    "relevance": 0.16,
    "specificity": 0.10,
    "directness": 0.15,
    "authenticity": 0.12,
    "procedural_integrity": 0.10,
    "internal_consistency": 0.10,
    "verifiability": 0.09,
}

SOURCE_TYPE_QUALITY = {
    "statement": 0.65,
    "evidence_image": 0.70,
    "report_image": 0.80,
    "official_report": 0.85,
    "forensic_report": 0.90,
    "party_submitted_image": 0.60,
    "manual_verified": 0.95,
}


def _load_models() -> list[dict]:
    models = []
    for path in sorted(MODEL_DIR.glob("*_v1.json")):
        model = json.loads(path.read_text(encoding="utf-8"))
        model["_path"] = path.relative_to(ROOT).as_posix()
        models.append(model)
    return models


def _start_rows() -> list[list[object]]:
    return [
        ["步骤", "要做什么", "非统计人员应如何操作", "为什么", "完成标志"],
        [1, "先定义统计对象", "一行只记录一个案件事件中的一个原子事实，例如 conduct 或 result_exists。不要把整份笔录当成一个真假标签。", "同一份材料可能包含多个事实，混在一起无法统计。", "已写明模型ID、节点、事件ID。"],
        [2, "确定样本范围", "写清起止日期、地区/单位、纳入条件和排除条件。一个月报警数据只能代表该月被纳入的报警样本。", "样本范围不同，计算出的比例不能直接互换。", "数据批次有唯一名称和口径说明。"],
        [3, "分别记录来源主张和独立真值", "来源说事实成立填“真”，来源否认填“假”；独立核验人员另填实际“真/假/未知”。", "来源说了什么与事实最终是否成立是两件事。", "原子事实记录的两个标签均由不同职责填写。"],
        [4, "未知就填未知", "没有足够材料时必须填“未知”，不能为了凑分母改成“假”。", "未知不是反证；错误处理会系统性压低参数。", "未知原因已填写，且不进入真/假分母。"],
        [5, "标记同源和复制", "同一视频截图、引用同一笔录的报告、转述链必须使用同一来源组ID。", "同源材料重复计数会制造虚假的多源印证。", "来源组ID和依赖记录均已复核。"],
        [6, "生成统计量", "根节点看真/假计数；来源表现看TP/FP/FN/TN；派生节点收集父节点值和子节点独立真值。", "三类参数需要不同的数据，不能用一个“真实性比例”替代。", "汇总表公式不再显示“先填数据”。"],
        [7, "只形成候选参数", "把计算值写入候选参数表，不直接覆盖JSON。逻辑回归的 intercept/weight 必须由离线脚本或统计人员拟合。", "一次手工改值无法保证模型整体稳定。", "候选参数有训练、验证样本量和验证结果。"],
        [8, "独立验证和审批", "用未参与拟合的数据回放，对比旧版；通过隐私、法律边界和人工审批后新建版本。", "同一批数据既拟合又验收会高估效果。", "模型发布记录为“已批准”，并记录参数哈希。"],
        ["核心区别", "Beta后验不是单案后验", "例如61/82表示该统计分组的长期比例估计，不等于新案件中某个事实有74.4%概率为真。", "单案后验还要结合当前案件其他独立证据和冻结的网络参数。", "使用者能复述这一差别。"],
        ["样本量提示", "模板阈值只是预警，不是发布保证", "有效样本<30只描述；任一真/假类别<10不拟合；达到100也仍需独立验证。逻辑回归还要看父节点数和类别不平衡。", "不存在适用于所有节点的统一最小样本量。", "统计复核人在发布记录中签署判断。"],
        ["禁止事项", "不能从真实数据自动学习", "法定年龄、金额、期限、证明标准、罪名、处罚、管辖、权威适用范围和网络安全边界不得由历史频率自动修改。", "这些是法律或治理规则，不是经验概率。", "不可学习项已逐项检查。"],
        ["隐私", "工作簿只放去标识化统计", "不得填写姓名、证件号、电话、地址、材料原文、图片或可识别的文件路径。", "统计模板不是案件卷宗。", "隐私检查为“通过”。"],
    ]


def _glossary_rows() -> list[list[object]]:
    return [
        ["名词", "白话解释", "本项目中的例子", "常见误解"],
        ["样本", "一条可独立核验的观察。", "某案件事件中的一个原子事实。", "一份笔录不一定只等于一个样本。"],
        ["真值", "独立核验后对事实成立与否的参考标签。", "真、假或未知。", "办案角色或模型输出不能自动充当真值。"],
        ["分子", "符合目标条件的数量。", "80条肯定主张中60条核验为真，分子是60。", "不能把未知计入分子。"],
        ["分母", "真正有资格参与比例计算的数量。", "60条正确肯定+20条错误肯定=80。", "未知、排除和同源副本不能混进分母。"],
        ["先验 prior", "没有当前案件软证据时，节点使用的基础值。", "conduct_result.conduct 当前为0.15。", "不是有罪率，也不是任何人的固定可信度。"],
        ["Beta后验", "把历史成功/失败计数与一个平滑先验合并后的比例估计。", "Alpha=1+60，Beta=1+20。", "这是群体参数后验，不是新案件事实后验。"],
        ["后验均值", "Beta分布中最常用的候选比例。", "61/(61+21)=0.7439。", "样本少时仍不能直接发布。"],
        ["95%区间", "参数仍可能落入的合理范围。区间越宽，说明数据越不稳定。", "由BETA.INV公式计算。", "不是说95%的案件一定在这个范围。"],
        ["TP", "来源肯定该事实，独立核验也为真。", "正确肯定。", "TP不是“这个人说真话”的人格评价。"],
        ["FP", "来源肯定该事实，但独立核验为假。", "错误肯定。", "不等于故意虚假陈述。"],
        ["FN", "来源否认该事实，但独立核验为真。", "错误否定。", "不等于主观恶意。"],
        ["TN", "来源否认该事实，独立核验也为假。", "正确否定。", "仍需确认来源是否独立。"],
        ["intercept", "所有父节点值为0时，子节点的基础对数优势。", "causation当前为-3.0。", "不能直接填写成百分比。"],
        ["weight", "某父节点升高时对子节点方向和强度的影响。", "alternative_cause权重为负。", "不能用“正确率74%”直接替换weight。"],
        ["校准", "让模型输出与独立验证数据的实际频率更一致。", "候选参数在独立验证集上比较。", "训练集表现好不等于已校准。"],
        ["单案后验", "冻结参数后，把某一案件的软证据送入网络得到的节点值。", "BayesianInferenceEngine.infer()输出。", "运行单案时不重新训练参数。"],
    ]


def _parameter_rows(models: list[dict]) -> list[list[object]]:
    rows = [[
        "参数层", "模型ID", "模型版本", "节点", "参数类型", "父节点/维度", "当前值",
        "能否根据真实数据修改", "必须收集什么", "如何构建统计量", "估计方法",
        "最低数据提示", "人工修改位置", "能否自动写回", "说明",
    ]]
    for model in models:
        for node in model["nodes"]:
            if node["type"] == "prior":
                rows.append([
                    "贝叶斯网络", model["model_id"], str(model["version"]), node["id"], "prior", "", node["prior"],
                    "可以，需验证", "该节点独立核验为真/假/未知的数量；明确样本范围和数据批次。",
                    "真数/(真数+假数)，未知不进分母；用Beta(Alpha,Beta)平滑。", "Beta-Binomial后验",
                    "有效样本<30仅描述；真或假任一<10不得发布。", f"{model['_path']} -> nodes[{node['id']}].prior", "否",
                    "只代表规定样本范围内的基础率；软证据存在时该值会被覆盖。",
                ])
            elif node["type"] == "logistic":
                parent_count = len(node["parents"])
                rows.append([
                    "贝叶斯网络", model["model_id"], str(model["version"]), node["id"], "intercept", "", node["intercept"],
                    "可以，但必须拟合", "每条样本的全部父节点值、子节点独立真值、样本权重和训练/验证划分。",
                    "一行一个事件；父节点0~1，子节点真/假；未知不进入拟合。", "带正则化的加权二项逻辑回归",
                    f"至少保留独立验证集；每个结果类别建议不少于10×{parent_count + 1}条，仅作预警。",
                    f"{model['_path']} -> nodes[{node['id']}].intercept", "否", "不能由单个比例手工换算。",
                ])
                for parent in node["parents"]:
                    rows.append([
                        "贝叶斯网络", model["model_id"], str(model["version"]), node["id"], "weight", parent, node["weights"][parent],
                        "可以，但必须拟合", "与intercept相同；还要检查父节点共线性、缺失和类别不平衡。",
                        "用所有父节点同时拟合，不能逐个父节点分别算比例后直接当权重。", "带正则化的加权二项逻辑回归",
                        f"至少保留独立验证集；每个结果类别建议不少于10×{parent_count + 1}条，仅作预警。",
                        f"{model['_path']} -> nodes[{node['id']}].weights.{parent}", "否", "方向变化必须由领域专家复核。",
                    ])

    rows.append([
        "主观证据", "subjective_evidence", "current", "all_claims", "base_rate_evidence", "", 2.0,
        "可以但不建议按月改", "不同证据量下的独立验证误差、覆盖率和不确定性表现。",
        "在留出验证集上比较不同W值的校准误差和弃权表现。", "网格搜索+独立验证",
        "至少跨多个时间批次验证。", "case_agent_demo/evidence_reasoning.py -> _BASE_RATE_EVIDENCE", "否", "它控制不确定性基线，不是事实先验。",
    ])
    for name, value in QUALITY_DIMENSION_WEIGHTS.items():
        rows.append([
            "证据质量", "evidence_quality", "current", "quality", "quality_weight", name, value,
            "可以但必须整体拟合", "该维度评分、独立真值、最终预测误差；需保留验证集。",
            "全部维度一起进入受约束模型，权重非负且总和为1。", "受约束回归/校准",
            "小样本时保持专家值，不逐项手改。", f"case_agent_demo/evidence_reasoning.py -> QUALITY_DIMENSION_WEIGHTS.{name}", "否", "不是按角色计算的真实性比例。",
        ])
    for name, value in SOURCE_TYPE_QUALITY.items():
        rows.append([
            "证据质量", "evidence_quality", "current", "source_type", "source_type_quality", name, value,
            "可监测，暂不直接改", "来源类别×谓词×立场的TP/FP/FN/TN、未知数和依赖组。",
            "先算肯定/否定主张正确率并做分层收缩，再评估是否需要重构全局默认值。", "分层Beta或校准模型",
            "单一月份、单一角色或单一谓词不得直接覆盖全局值。", f"case_agent_demo/evidence_reasoning.py -> SOURCE_TYPE_QUALITY.{name}", "否", "角色只能作为分组变量，不能形成永久人格可信度。",
        ])

    authority = json.loads(AUTHORITY_RULES.read_text(encoding="utf-8"))
    for rule in authority.get("rules", []):
        for field in ("mean", "strength"):
            rows.append([
                "权威锚定", "authority", "current", rule["id"], field, "", rule[field],
                "只监测，不自动修改", "有效文书数、被推翻/替代数、程序无效数及选择机制说明。",
                "只形成监测率；需考虑重新鉴定样本的选择偏差。", "治理复核+敏感性分析",
                "必须由法律/专业委员会批准。", f"config/authority_rules.json -> rules[{rule['id']}].{field}", "否", "权威范围和有效性不是普通来源正确率。",
            ])

    rows.extend([
        ["网络结构", "registry", "current", "all", "edge/required_inputs", "", "见registry.json", "禁止自动修改", "只能收集误差案例作为专家审查材料。", "不从频率自动生成边。", "因果/领域专家审查", "无自动样本阈值。", "config/bayesian_models/registry.json", "否", "边方向和必需输入属于模型设计与安全边界。"],
        ["法律规则", "legal_rules", "current", "all", "threshold/rule", "", "见法律配置", "禁止", "不采集训练统计。", "确定性法律规则。", "人工法律审查", "不适用。", "law_DB / legal_knowledge / legal element config", "否", "法定年龄、数额、期限、证明标准、罪名和处罚不得由历史频率学习。"],
    ])
    return rows


def _atomic_rows() -> list[list[object]]:
    return [[
        "记录ID", "样本ID", "数据批次", "观察日期", "模型ID", "模型版本", "节点/谓词", "事件ID(匿名)",
        "来源类别", "来源角色", "来源组ID", "来源主张标签", "独立核验标签", "核验依据类别",
        "核验人员代号", "复核状态", "是否纳入统计", "统计分类", "排除/未知原因", "抽取器版本", "Prompt版本", "备注",
    ]]


def _root_rows(models: list[dict]) -> list[list[object]]:
    rows = [[
        "模型ID", "模型版本", "根节点", "数据批次", "当前prior", "真数", "假数", "未知数", "有效样本数",
        "先验Alpha", "先验Beta", "后验Alpha", "后验Beta", "后验均值", "95%下限", "95%上限", "样本提示", "候选决策", "备注",
    ]]
    for model in models:
        for node in model["nodes"]:
            if node["type"] == "prior":
                rows.append([
                    model["model_id"], str(model["version"]), node["id"], "", node["prior"],
                    None, None, None, None, 1.0, 1.0, None, None, None, None, None, None, "不采用", "",
                ])
    return rows


def _source_summary_rows() -> list[list[object]]:
    return [[
        "模型ID", "模型版本", "节点/谓词", "数据批次", "来源类别", "来源角色", "指标类型",
        "TP", "FP", "FN", "TN", "未知数", "先验Alpha", "先验Beta", "后验Alpha", "后验Beta",
        "后验均值", "95%下限", "95%上限", "有效分母", "样本提示", "候选用途", "备注",
    ]]


def _cpd_sample_rows() -> list[list[object]]:
    return [[
        "记录ID", "样本ID", "数据批次", "模型ID", "模型版本", "子节点",
        "父节点1", "值1", "父节点2", "值2", "父节点3", "值3", "父节点4", "值4", "父节点5", "值5",
        "子节点独立核验标签", "样本权重", "数据用途", "核验人员代号", "复核状态", "是否纳入统计", "备注",
    ]]


def _cpd_summary_rows() -> list[list[object]]:
    return [[
        "模型ID", "模型版本", "子节点", "父节点状态组合", "数据批次", "子节点真数", "子节点假数", "未知数",
        "先验Alpha", "先验Beta", "后验Alpha", "后验Beta", "后验均值", "95%下限", "95%上限", "样本提示", "说明",
    ]]


def _extraction_rows() -> list[list[object]]:
    return [[
        "抽取器ID", "模型/Prompt版本", "字段或谓词", "数据批次", "TP", "FP", "FN", "TN", "未知数",
        "精确率", "召回率", "F1", "有效样本数", "样本提示", "说明",
    ]]


def _dependency_rows() -> list[list[object]]:
    return [[
        "依赖记录ID", "数据批次", "样本ID", "来源A类别", "来源B类别", "共同来源组ID", "依赖类型",
        "依赖复核结论", "处理方式", "复核人员代号", "复核日期", "备注",
    ]]


def _authority_rows() -> list[list[object]]:
    return [[
        "规则ID", "文书类型", "适用谓词", "数据批次", "复核总数", "结论维持数", "被推翻/替代数", "程序无效数", "未知数",
        "监测维持率", "选择偏差说明", "权威状态", "复核人员代号", "复核日期", "能否自动改参数", "备注",
    ]]


def _candidate_rows(models: list[dict]) -> list[list[object]]:
    rows = [[
        "候选ID", "模型ID", "当前版本", "目标版本", "节点", "参数类型", "父节点/维度", "当前值", "候选值",
        "估计方法", "训练有效样本", "验证有效样本", "验证指标", "与旧版比较", "参数方向复核", "独立复核状态",
        "审批状态", "修改理由", "目标文件/字段", "是否已写入新版本", "备注",
    ]]
    index = 1
    for row in _parameter_rows(models)[1:]:
        if row[0] != "贝叶斯网络":
            continue
        rows.append([
            f"CAND-{index:03d}", row[1], row[2], "", row[3], row[4], row[5], row[6], "", row[10],
            "", "", "", "", "待复核", "未开始", "草稿", "", row[12], "否", "",
        ])
        index += 1
    return rows


def _release_rows() -> list[list[object]]:
    return [[
        "模型ID", "模型版本", "参数哈希", "校准状态", "数据截止日期", "数据批次", "训练样本数", "验证样本数",
        "独立复核状态", "离线审批状态", "审批角色代号", "审批日期", "隐私检查", "法律边界检查", "回放结果",
        "发布包相对路径", "替代版本", "回滚包路径", "备注",
    ]]


def _nonlearnable_rows() -> list[list[object]]:
    return [
        ["项目", "为什么不能从历史频率自动学习", "允许做什么", "发布前检查"],
        ["法定年龄、金额、次数和期限", "它们来自现行法律，不是经验概率。", "版本化录入法律规则并核对生效时间。", "候选参数中不得出现。"],
        ["罪名、责任、处罚和证明标准", "历史办案结果不能替代法律判断。", "由法律规则与人工审查处理。", "模型输出不得命名为有罪概率。"],
        ["管辖、身份、职权、授权、抗辩和例外", "属于规范适用条件。", "保存结构化事实，交由规则引擎/人工判断。", "不得写入prior或weight。"],
        ["网络边方向和required_inputs", "相关性不等于因果关系；删掉必需输入可能越过安全边界。", "用误差案例支持专家评审，新建模型版本。", "registry变更必须单独评审。"],
        ["权威材料适用范围", "文书是否能证明某个事实由专业和法律范围决定。", "监测被推翻率，检查程序和版本。", "不得只按历史维持率自动扩大范围。"],
        ["按报警人/嫌疑人角色设永久可信度", "角色不是人格真值，且不同谓词、场景和核验条件差异很大。", "角色只作为分层变量，结合谓词、立场、来源依赖和时间批次统计。", "任何全局角色分数不得发布。"],
        ["同源材料重复计数", "副本不提供新的独立信息。", "使用来源组去重或经批准的依赖模型。", "同一origin_evidence只计一次。"],
    ]


def _example_rows() -> list[list[object]]:
    return [
        ["示例", "输入1", "输入2", "输入3", "先验Alpha", "先验Beta", "后验Alpha", "后验Beta", "后验均值", "如何解释"],
        ["报警人对某一谓词的肯定主张正确率（示例，不代表真实数据）", 60, 20, 10, 1, 1, "=E2+B2", "=F2+C2", "=G2/(G2+H2)", "60条正确肯定、20条错误肯定、10条未知。未知不进分母；结果只描述该谓词、该批次、该核验口径。"],
        ["某根节点prior（示例，不代表真实数据）", 30, 50, 20, 1, 1, "=E3+B3", "=F3+C3", "=G3/(G3+H3)", "30真、50假、20未知；Beta平滑后的候选prior。只有样本框具有代表性时才有意义。"],
        ["逻辑回归weight", "父节点逐行值", "子节点独立真值", "训练/验证划分", "-", "-", "-", "-", "不能用简单比例", "把CPD样本记录导出，由离线脚本同时拟合intercept和全部weights；人工不能把74%直接写成0.74权重。"],
        ["单案后验", "冻结参数", "当前案件软证据", "网络推理", "-", "-", "-", "-", "由程序计算", "运行案件时不修改参数。它与上面用于估计群体参数的Beta后验不是同一个概念。"],
    ]


def _manual_update_rows() -> list[list[object]]:
    return [
        ["顺序", "人工操作", "检查点", "错误做法"],
        [1, "冻结数据批次，确认未知、排除和同源记录已处理。", "批次口径、起止日期、纳入/排除条件已记录。", "边收集边覆盖线上参数。"],
        [2, "在根节点统计、来源表现统计和CPD样本记录中完成汇总。", "分母不含未知；同源只计一次。", "把未知当假，或把十张同源截图计十次。"],
        [3, "root prior使用Beta后验形成候选；intercept/weight由离线逻辑回归形成候选。", "候选值、训练样本量、方法写入候选参数。", "用角色正确率直接替换weight。"],
        [4, "复制原模型JSON为新版本文件，不覆盖旧文件。", "目标版本、目标文件、替代版本明确。", "直接修改当前已发布JSON。"],
        [5, "只写入已批准字段：prior、intercept或对应weights[parent]。", "数值范围、方向和JSON结构通过校验。", "修改registry边、required_inputs或法律门槛。"],
        [6, "更新version和calibration_status，运行模型校验、全量测试和六类回放。", "训练集外的验证结果优于或不劣于旧版。", "只看训练数据结果。"],
        [7, "记录参数哈希、审批、隐私检查、发布包和回滚包。", "模型发布记录所有必填项完成。", "没有回滚包就发布。"],
        [8, "部署后只监测，不在线学习；达到下一批次后再次离线校准。", "单案运行不会改变参数。", "每处理一个案件自动更新参数。"],
    ]


def _add_sheet(workbook: Workbook, title: str, rows: list[list[object]]) -> None:
    sheet = workbook.create_sheet(title)
    for row in rows:
        sheet.append(row)


def _add_enumerations(workbook: Workbook) -> None:
    sheet = workbook.create_sheet("枚举值")
    for column, (range_name, (header, values)) in enumerate(ENUMS.items(), start=1):
        sheet.cell(1, column, header)
        for row, value in enumerate(values, start=2):
            sheet.cell(row, column, value)
        reference = (
            f"{quote_sheetname(sheet.title)}!${get_column_letter(column)}$2:"
            f"${get_column_letter(column)}${len(values) + 1}"
        )
        workbook.defined_names.add(DefinedName(range_name, attr_text=reference))


def _list_validation(sheet, cell_range: str, range_name: str, prompt: str) -> None:
    validation = DataValidation(type="list", formula1=f"={range_name}", allow_blank=True)
    validation.error = "请从下拉列表选择，不要自行创造近义值。"
    validation.errorTitle = "无效值"
    validation.prompt = prompt
    validation.promptTitle = "填写提示"
    validation.showErrorMessage = True
    validation.showInputMessage = True
    sheet.add_data_validation(validation)
    validation.add(cell_range)


def _custom_validation(sheet, cell_range: str, formula: str, prompt: str) -> None:
    validation = DataValidation(type="custom", formula1=formula, allow_blank=True)
    validation.error = "请输入0到1之间的数，或填写“未知”。"
    validation.errorTitle = "无效概率值"
    validation.prompt = prompt
    validation.promptTitle = "父节点软证据"
    validation.showErrorMessage = True
    validation.showInputMessage = True
    sheet.add_data_validation(validation)
    validation.add(cell_range)


def _populate_formulas(workbook: Workbook) -> None:
    atomic = workbook["原子事实记录"]
    for row in range(2, ATOMIC_LAST_ROW + 1):
        atomic.cell(row, 18, (
            f'=IF(Q{row}<>"是","排除",IF(OR(L{row}="",M{row}=""),"",'
            f'IF(OR(L{row}="未知",M{row}="未知"),"未知",'
            f'IF(AND(L{row}="真",M{row}="真"),"TP",'
            f'IF(AND(L{row}="真",M{row}="假"),"FP",'
            f'IF(AND(L{row}="假",M{row}="真"),"FN","TN"))))))'
        ))

    atom = quote_sheetname("原子事实记录")
    root = workbook["根节点统计"]
    for row in range(2, root.max_row + 1):
        criteria = (
            f"{atom}!$C$2:$C${ATOMIC_LAST_ROW},$D{row},"
            f"{atom}!$E$2:$E${ATOMIC_LAST_ROW},$A{row},"
            f"{atom}!$F$2:$F${ATOMIC_LAST_ROW},$B{row},"
            f"{atom}!$G$2:$G${ATOMIC_LAST_ROW},$C{row},"
            f"{atom}!$Q$2:$Q${ATOMIC_LAST_ROW},\"是\""
        )
        root.cell(row, 6, f'=IF($D{row}="","",COUNTIFS({criteria},{atom}!$M$2:$M${ATOMIC_LAST_ROW},"真"))')
        root.cell(row, 7, f'=IF($D{row}="","",COUNTIFS({criteria},{atom}!$M$2:$M${ATOMIC_LAST_ROW},"假"))')
        root.cell(row, 8, f'=IF($D{row}="","",COUNTIFS({criteria},{atom}!$M$2:$M${ATOMIC_LAST_ROW},"未知"))')
        root.cell(row, 9, f'=IF($D{row}="","",F{row}+G{row})')
        root.cell(row, 12, f'=IF(OR(J{row}="",F{row}=""),"",J{row}+F{row})')
        root.cell(row, 13, f'=IF(OR(K{row}="",G{row}=""),"",K{row}+G{row})')
        root.cell(row, 14, f'=IF(OR(L{row}="",M{row}=""),"",L{row}/(L{row}+M{row}))')
        root.cell(row, 15, f'=IF(N{row}="","",BETA.INV(0.025,L{row},M{row}))')
        root.cell(row, 16, f'=IF(N{row}="","",BETA.INV(0.975,L{row},M{row}))')
        root.cell(row, 17, (
            f'=IF(D{row}="","先填数据批次",IF(I{row}<30,"样本太少：仅描述",'
            f'IF(MIN(F{row},G{row})<10,"类别失衡：不可发布",'
            f'IF(I{row}<100,"试点：需独立验证","可进入独立验证"))))'
        ))

    source = workbook["来源表现统计"]
    for row in range(2, SUMMARY_LAST_ROW + 1):
        base = (
            f"{atom}!$C$2:$C${ATOMIC_LAST_ROW},$D{row},"
            f"{atom}!$E$2:$E${ATOMIC_LAST_ROW},$A{row},"
            f"{atom}!$F$2:$F${ATOMIC_LAST_ROW},$B{row},"
            f"{atom}!$G$2:$G${ATOMIC_LAST_ROW},$C{row},"
            f"{atom}!$I$2:$I${ATOMIC_LAST_ROW},$E{row},"
            f"{atom}!$J$2:$J${ATOMIC_LAST_ROW},$F{row},"
            f"{atom}!$Q$2:$Q${ATOMIC_LAST_ROW},\"是\""
        )
        guard = f'IF(COUNTA(A{row}:G{row})<7,"",'
        for column, category in zip(range(8, 12), ("TP", "FP", "FN", "TN")):
            source.cell(row, column, f'={guard}COUNTIFS({base},{atom}!$R$2:$R${ATOMIC_LAST_ROW},"{category}"))')
        source.cell(row, 12, f'={guard}COUNTIFS({base},{atom}!$R$2:$R${ATOMIC_LAST_ROW},"未知"))')
        source.cell(row, 15, (
            f'=IF(OR(G{row}="",M{row}="",N{row}=""),"",M{row}+'
            f'IF(G{row}="肯定主张正确率",H{row},IF(G{row}="否定主张正确率",K{row},IF(G{row}="灵敏度",H{row},K{row}))))'
        ))
        source.cell(row, 16, (
            f'=IF(OR(G{row}="",M{row}="",N{row}=""),"",N{row}+'
            f'IF(G{row}="肯定主张正确率",I{row},IF(G{row}="否定主张正确率",J{row},IF(G{row}="灵敏度",J{row},I{row}))))'
        ))
        source.cell(row, 17, f'=IF(OR(O{row}="",P{row}=""),"",O{row}/(O{row}+P{row}))')
        source.cell(row, 18, f'=IF(Q{row}="","",BETA.INV(0.025,O{row},P{row}))')
        source.cell(row, 19, f'=IF(Q{row}="","",BETA.INV(0.975,O{row},P{row}))')
        source.cell(row, 20, f'=IF(OR(O{row}="",P{row}=""),"",O{row}+P{row}-M{row}-N{row})')
        source.cell(row, 21, (
            f'=IF(T{row}="","先填完整分组",IF(T{row}<30,"样本太少：仅描述",'
            f'IF(T{row}<100,"只作试点候选","可进入独立验证")))'
        ))

    cpd = workbook["CPD组合统计"]
    for row in range(2, SUMMARY_LAST_ROW + 1):
        cpd.cell(row, 11, f'=IF(OR(I{row}="",F{row}=""),"",I{row}+F{row})')
        cpd.cell(row, 12, f'=IF(OR(J{row}="",G{row}=""),"",J{row}+G{row})')
        cpd.cell(row, 13, f'=IF(OR(K{row}="",L{row}=""),"",K{row}/(K{row}+L{row}))')
        cpd.cell(row, 14, f'=IF(M{row}="","",BETA.INV(0.025,K{row},L{row}))')
        cpd.cell(row, 15, f'=IF(M{row}="","",BETA.INV(0.975,K{row},L{row}))')
        cpd.cell(row, 16, (
            f'=IF(F{row}+G{row}=0,"先填真/假计数",IF(F{row}+G{row}<30,"组合样本太少",'
            f'IF(MIN(F{row},G{row})<5,"结果类别不足","可作CPT参考；逻辑权重仍需整体拟合")))'
        ))

    extraction = workbook["抽取质量统计"]
    for row in range(2, SUMMARY_LAST_ROW + 1):
        extraction.cell(row, 10, f'=IF(E{row}+F{row}=0,"",E{row}/(E{row}+F{row}))')
        extraction.cell(row, 11, f'=IF(E{row}+G{row}=0,"",E{row}/(E{row}+G{row}))')
        extraction.cell(row, 12, f'=IF(OR(J{row}="",K{row}="",J{row}+K{row}=0),"",2*J{row}*K{row}/(J{row}+K{row}))')
        extraction.cell(row, 13, f'=IF(COUNTA(E{row}:H{row})=0,"",SUM(E{row}:H{row}))')
        extraction.cell(row, 14, f'=IF(M{row}="","先填计数",IF(M{row}<30,"样本太少：仅描述",IF(M{row}<100,"试点","可进入验证")))')

    authority = workbook["权威锚定复核"]
    for row in range(2, SUMMARY_LAST_ROW + 1):
        authority.cell(row, 10, f'=IF(F{row}+G{row}=0,"",F{row}/(F{row}+G{row}))')


def _ensure_template_rows(workbook: Workbook) -> None:
    for sheet_name, last_row in {
        "CPD样本记录": CPD_LAST_ROW,
        "来源依赖": SUMMARY_LAST_ROW,
        "模型发布记录": SUMMARY_LAST_ROW,
    }.items():
        sheet = workbook[sheet_name]
        sheet.cell(last_row, sheet.max_column, "")


def _add_validations(workbook: Workbook) -> None:
    atomic = workbook["原子事实记录"]
    _list_validation(atomic, f"I2:I{ATOMIC_LAST_ROW}", "SourceTypes", "选择材料的结构化来源类别。")
    _list_validation(atomic, f"J2:J{ATOMIC_LAST_ROW}", "SourceRoles", "角色只用于分层，不能直接变成可信度。")
    _list_validation(atomic, f"L2:M{ATOMIC_LAST_ROW}", "TruthLabels", "未知不得填成假。")
    _list_validation(atomic, f"P2:P{ATOMIC_LAST_ROW}", "ReviewStatuses", "有分歧时必须保留并裁决。")
    _list_validation(atomic, f"Q2:Q{ATOMIC_LAST_ROW}", "YesNoValues", "只有完成独立核验且不重复的记录才纳入。")

    source = workbook["来源表现统计"]
    _list_validation(source, f"E2:E{SUMMARY_LAST_ROW}", "SourceTypes", "来源类别必须与原子记录一致。")
    _list_validation(source, f"F2:F{SUMMARY_LAST_ROW}", "SourceRoles", "必须按谓词和立场一起解释。")
    _list_validation(source, f"G2:G{SUMMARY_LAST_ROW}", "SourceMetrics", "优先查看肯定/否定主张正确率。")
    _list_validation(source, f"V2:V{SUMMARY_LAST_ROW}", "ParameterDecisions", "统计结果先作为候选。")

    root = workbook["根节点统计"]
    _list_validation(root, f"R2:R{root.max_row}", "ParameterDecisions", "不得直接写回模型。")

    cpd_samples = workbook["CPD样本记录"]
    for column in ("H", "J", "L", "N", "P"):
        _custom_validation(
            cpd_samples,
            f"{column}2:{column}{CPD_LAST_ROW}",
            f'=OR({column}2="",{column}2="未知",AND(ISNUMBER({column}2),{column}2>=0,{column}2<=1))',
            "输入0到1的软证据值；无法核验时填“未知”。",
        )
    _list_validation(cpd_samples, f"Q2:Q{CPD_LAST_ROW}", "TruthLabels", "子节点必须由独立核验给出真/假/未知。")
    _list_validation(cpd_samples, f"S2:S{CPD_LAST_ROW}", "DataSplits", "验证和测试数据不得参与拟合。")
    _list_validation(cpd_samples, f"U2:U{CPD_LAST_ROW}", "ReviewStatuses", "分歧必须裁决。")
    _list_validation(cpd_samples, f"V2:V{CPD_LAST_ROW}", "YesNoValues", "未知子节点不纳入拟合。")

    authority = workbook["权威锚定复核"]
    _list_validation(authority, f"L2:L{SUMMARY_LAST_ROW}", "AuthorityStatuses", "权威范围由专业和法律复核确定。")

    candidate = workbook["候选参数"]
    _list_validation(candidate, f"P2:P{candidate.max_row}", "ReviewStatuses", "候选参数必须独立复核。")
    _list_validation(candidate, f"Q2:Q{candidate.max_row}", "ApprovalStatuses", "只有已批准的候选可写入新版本。")
    _list_validation(candidate, f"T2:T{candidate.max_row}", "YesNoValues", "只能写入新版本文件。")

    release = workbook["模型发布记录"]
    _list_validation(release, f"D2:D{SUMMARY_LAST_ROW}", "CalibrationStatuses", "没有真实独立验证不得标validated。")
    _list_validation(release, f"I2:I{SUMMARY_LAST_ROW}", "ReviewStatuses", "独立复核必须完成。")
    _list_validation(release, f"J2:J{SUMMARY_LAST_ROW}", "ApprovalStatuses", "必须离线审批。")
    _list_validation(release, f"M2:M{SUMMARY_LAST_ROW}", "PrivacyStatuses", "隐私检查不通过不得发布。")


def _style_sheet(sheet) -> None:
    sheet.freeze_panes = "A2"
    sheet.sheet_view.showGridLines = False
    sheet.sheet_view.zoomScale = 85
    sheet.auto_filter.ref = f"A1:{get_column_letter(sheet.max_column)}{max(1, sheet.max_row)}"
    sheet.row_dimensions[1].height = 34
    for cell in sheet[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = THIN_BORDER
        cell.protection = Protection(locked=True)
    for row in sheet.iter_rows(min_row=2):
        for cell in row:
            cell.font = BODY_FONT
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            if cell.data_type == "f":
                cell.fill = FORMULA_FILL
    for column in range(1, sheet.max_column + 1):
        values = [str(sheet.cell(row, column).value or "") for row in range(1, min(sheet.max_row, 80) + 1)]
        width = min(max(max(map(len, values), default=8) + 2, 12), 38)
        sheet.column_dimensions[get_column_letter(column)].width = width


def _style_workbook(workbook: Workbook) -> None:
    for sheet in workbook.worksheets:
        _style_sheet(sheet)

    for name in ("开始这里", "名词解释", "不可学习项", "人工修改流程", "完整示例"):
        workbook[name].sheet_properties.tabColor = "20639B"
    for name in ("原子事实记录", "CPD样本记录", "来源依赖"):
        workbook[name].sheet_properties.tabColor = "F6B26B"
    for name in ("根节点统计", "来源表现统计", "CPD组合统计", "抽取质量统计"):
        workbook[name].sheet_properties.tabColor = "6AA84F"
    for name in ("候选参数", "模型发布记录", "权威锚定复核"):
        workbook[name].sheet_properties.tabColor = "674EA7"
    workbook["枚举值"].sheet_properties.tabColor = "999999"

    workbook["参数总览"].freeze_panes = "A2"
    workbook["参数总览"].auto_filter.ref = f"A1:O{workbook['参数总览'].max_row}"

    input_ranges = {
        "原子事实记录": ["A:Q", "S:V"],
        "根节点统计": ["D:D", "J:K", "R:S"],
        "来源表现统计": ["A:G", "M:N", "V:W"],
        "CPD样本记录": ["A:W"],
        "CPD组合统计": ["A:J", "Q:Q"],
        "抽取质量统计": ["A:I", "O:O"],
        "来源依赖": ["A:L"],
        "权威锚定复核": ["A:I", "K:P"],
        "候选参数": ["D:D", "I:U"],
        "模型发布记录": ["A:S"],
    }
    for sheet_name, ranges in input_ranges.items():
        sheet = workbook[sheet_name]
        for range_ref in ranges:
            for row in sheet[range_ref]:
                cells = row if isinstance(row, tuple) else (row,)
                for cell in cells:
                    if cell.row > 1 and cell.data_type != "f":
                        cell.fill = INPUT_FILL

    warning_columns = {
        "根节点统计": "Q",
        "来源表现统计": "U",
        "CPD组合统计": "P",
    }
    for sheet_name, column in warning_columns.items():
        sheet = workbook[sheet_name]
        sheet.conditional_formatting.add(
            f"{column}2:{column}{sheet.max_row}",
            FormulaRule(
                formula=[
                    f'OR(ISNUMBER(SEARCH("样本",{column}2)),ISNUMBER(SEARCH("不可",{column}2)))'
                ],
                fill=WARNING_FILL,
            ),
        )

    workbook["原子事实记录"].conditional_formatting.add(
        f"R2:R{ATOMIC_LAST_ROW}", FormulaRule(formula=['R2="未知"'], fill=WARNING_FILL)
    )
    workbook["原子事实记录"].conditional_formatting.add(
        f"R2:R{ATOMIC_LAST_ROW}", FormulaRule(formula=['R2="排除"'], fill=CURRENT_FILL)
    )
    workbook["模型发布记录"].conditional_formatting.add(
        f"M2:M{SUMMARY_LAST_ROW}", CellIsRule(operator="equal", formula=['"不通过"'], fill=DANGER_FILL)
    )

    probability_columns = {
        "根节点统计": ("E", "N", "O", "P"),
        "来源表现统计": ("Q", "R", "S"),
        "CPD组合统计": ("M", "N", "O"),
        "抽取质量统计": ("J", "K", "L"),
        "权威锚定复核": ("J",),
        "完整示例": ("I",),
    }
    for sheet_name, columns in probability_columns.items():
        sheet = workbook[sheet_name]
        for column in columns:
            for cell in sheet[column][1:]:
                cell.number_format = "0.0000"

    for sheet_name in ("原子事实记录", "来源依赖", "权威锚定复核", "模型发布记录"):
        sheet = workbook[sheet_name]
        for row in sheet.iter_rows(min_row=2):
            for cell in row:
                if "日期" in str(sheet.cell(1, cell.column).value or ""):
                    cell.number_format = "yyyy-mm-dd"

    workbook["开始这里"]["A1"].comment = Comment(
        "从本页按1到8执行。工作簿不会自动修改项目参数。", "Va1ha11a_demo"
    )
    workbook["参数总览"]["H1"].comment = Comment(
        "“可以”仍表示先形成候选、独立验证和审批，不代表可以直接覆盖配置。", "Va1ha11a_demo"
    )


def _normalize_archive(path: Path) -> None:
    normalized = path.with_name(f".{path.name}.tmp")
    with ZipFile(path) as source, ZipFile(normalized, "w", compression=ZIP_DEFLATED, compresslevel=9) as target:
        for info in source.infolist():
            info.date_time = (1980, 1, 1, 0, 0, 0)
            info.create_system = 0
            data = source.read(info.filename)
            if info.filename == "docProps/core.xml":
                root = ElementTree.fromstring(data)
                modified = root.find(f"{{{CORE_NAMESPACES['dcterms']}}}modified")
                if modified is not None:
                    modified.text = "2026-07-13T00:00:00Z"
                data = ElementTree.tostring(root, encoding="utf-8")
            target.writestr(info, data)
    normalized.replace(path)


def build_workbook(output_path: str | Path = DEFAULT_OUTPUT) -> Path:
    output = Path(output_path)
    output.parent.mkdir(parents=True, exist_ok=True)
    models = _load_models()

    workbook = Workbook()
    workbook.remove(workbook.active)
    workbook.properties.creator = "Va1ha11a_demo"
    workbook.properties.title = "贝叶斯参数采集与人工更新模板"
    workbook.properties.subject = "真实数据采集、统计构建、候选参数和离线审批"
    workbook.properties.created = datetime(2026, 7, 13, tzinfo=timezone.utc)
    workbook.properties.modified = datetime(2026, 7, 13, tzinfo=timezone.utc)
    workbook.calculation.fullCalcOnLoad = True
    workbook.calculation.forceFullCalc = True
    workbook.calculation.calcMode = "auto"

    sheets = [
        ("开始这里", _start_rows()),
        ("名词解释", _glossary_rows()),
        ("参数总览", _parameter_rows(models)),
        ("原子事实记录", _atomic_rows()),
        ("根节点统计", _root_rows(models)),
        ("来源表现统计", _source_summary_rows()),
        ("CPD样本记录", _cpd_sample_rows()),
        ("CPD组合统计", _cpd_summary_rows()),
        ("抽取质量统计", _extraction_rows()),
        ("来源依赖", _dependency_rows()),
        ("权威锚定复核", _authority_rows()),
        ("候选参数", _candidate_rows(models)),
        ("模型发布记录", _release_rows()),
        ("不可学习项", _nonlearnable_rows()),
        ("完整示例", _example_rows()),
        ("人工修改流程", _manual_update_rows()),
    ]
    for title, rows in sheets:
        _add_sheet(workbook, title, rows)
    _add_enumerations(workbook)
    _ensure_template_rows(workbook)
    _populate_formulas(workbook)
    _add_validations(workbook)
    _style_workbook(workbook)

    workbook.save(output)
    _normalize_archive(output)
    return output


if __name__ == "__main__":
    print(build_workbook())
